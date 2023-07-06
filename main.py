from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/', methods=['POST'])
def generate_id():
    file = request.files['file']
    if file:
        file.save('data/ean.xlsx')

        tabela_ean = pd.read_excel('data/ean.xlsx')
        tabela_id = pd.read_excel('data/id.xlsx')

        rezultujuca_tabela = pd.merge(tabela_ean, tabela_id, on='EAN', how='left')
        rezultujuca_tabela['ID'].fillna('Nemam ID', inplace=True)
        rezultujuca_tabela.loc[rezultujuca_tabela['EAN'].isnull(), 'ID'] = ''

        book = Workbook()
        sheet = book.active

        header = list(rezultujuca_tabela.columns)
        sheet.append(header)
        for _, row in rezultujuca_tabela.iterrows():
            sheet.append(list(row))

        for cell in sheet['A'][1:]:
            cell.number_format = numbers.FORMAT_NUMBER

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        book.save('data/rezultat.xlsx')
        return render_template('download.html')
    else:
        return "Nije otpremljen fajl."

@app.route('/download')
def download():
    return send_file('data/rezultat.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run()
